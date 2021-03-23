from docxtpl import DocxTemplate
import time
import openpyxl
import datetime

# t = (2018, 9, 13, 11, 35, 45, 1, 48, 0)


n = 64
day = time.localtime()
d = day[2]
# print(type(d))
les = ['Проектирование и разработка робототехнических систем', "Программное обеспечение роботов", "Информатика"]

themes_list = []
date_dict = []
lessons = {}
number_list = []


def table(group, lesson, ):
    pattern = f'{lesson}{group}'

    wb = openpyxl.load_workbook('Робототехника 7-8.xlsx', data_only=True)
    # печатаем список листов
    sheets = wb.sheetnames
    for sheet in sheets:
        print(sheet)

    sheet = wb.active
    print(sheet)
    print(sheet['A1'])
    print(sheet['A1'].value)
    # Перебираем темы
    for row in sheet['B54':'B70']:
        for cellObj in row:
            if cellObj.value == None or cellObj.value == " ":
                continue
            # print(cellObj.value)
            themes_list.append(cellObj.value)
    # Перебираем номера уроков
    for row in sheet['A54':'A70']:
        for cellObj in row:
            if cellObj.value == None or cellObj.value == " ":
                continue
            # print(cellObj.value)
            number_list.append(cellObj.value)
    # Перебираем даты
    for row in sheet['C54':'C70']:
        for cellObj in row:
            if cellObj.value == None or cellObj.value == " ":
                continue
            print(type(cellObj.value))

            if cellObj.value.month == 1:
                result_mounth = 'января'
            elif cellObj.value.month == 2:
                result_mounth = 'февраля'
            elif cellObj.value.month == 3:
                result_mounth = 'марта'
            elif cellObj.value.month == 4:
                result_mounth = 'апреля'
            elif cellObj.value.month == 5:
                result_mounth = 'мая'
            elif cellObj.value.month == 6:
                result_mounth = 'июня'
            elif cellObj.value.month == 7:
                result_mounth = 'июля'
            elif cellObj.value.month == 8:
                result_mounth = 'августа'
            elif cellObj.value.month == 9:
                result_mounth = 'сентября'
            elif cellObj.value.month == 10:
                result_mounth = 'октября'
            elif cellObj.value.month == 11:
                result_mounth = 'ноября'
            elif cellObj.value.month == 12:
                result_mounth = 'декабря'
            date_dict.append(f"{cellObj.value.day} {result_mounth} {cellObj.value.year}")
    lessons = {x: y for x in themes_list for y in date_dict}
    print(date_dict)

    # cell_value = sheet.cell(column,row).value  # cell_value = sheet.cell(строка, столбец)
    # print(cell_value)
    print(lessons)


def manual():
    context = {'name': 'Молотков М.А.', 'lesson': 'Программное обеспечение роботов',
               'theme': 'Обработка последовательностей с помощью свёрточных нейронных сетей',
               'date': f'18 марта 2021 г', 'number': f'64', 'class': '9-10'}
    doc.render(context)
    doc.save(f"{context['class']}/{context['lesson']}/Занятие_64.docx")


# manual()

def automatic(pattern, group, lesson, hours):
    for x in range(hours):
        doc = DocxTemplate(pattern)
        day = time.localtime()
        d = day[2]
        n = 64
        context = {'name': 'Молотков М.А.', 'lesson': lesson,
                   'theme': themes_list[x], 'date': date_dict[x], 'number': number_list[x], 'class': group}

        doc.render(context)
        doc.save(f"{context['class']}/{context['lesson']}/Занятие_{number_list[x]}.docx")

        print(context['date'])
        print(context['number'])
        n += 1
        if x % 2 == 0:
            d += 5
        else:
            d += 2


pattern = "допы.docx"
group = '7-8'
lesson = 'Проектирование и разработка робототехнических систем'
hours = 70-54
table(pattern, group)
automatic(pattern, group, lesson, hours)
